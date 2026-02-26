from pathlib import Path

from openpyxl import Workbook, load_workbook


def read_excel(path: str | Path) -> list[dict]:
    """读取 Excel 文件，返回字典列表（首行为表头）。"""
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < 2:
        return []
    headers = [str(h) for h in rows[0]]
    return [dict(zip(headers, row)) for row in rows[1:]]


def write_excel(path: str | Path, data: list[dict]) -> None:
    """将字典列表写入 Excel 文件（首行为表头）。"""
    if not data:
        return
    wb = Workbook()
    ws = wb.active
    headers = list(data[0].keys())
    ws.append(headers)
    for row in data:
        ws.append([row.get(h) for h in headers])
    wb.save(path)
