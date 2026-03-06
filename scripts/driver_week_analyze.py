"""
司机周数据分析脚本 — 整合自 func/driver_week_analyze/
支持 run_dwa（生成DWA分析）和 run_comparison（生成周对周对比报表）。
"""

from __future__ import annotations

import math
import re
from pathlib import Path
from typing import Callable

import pandas as pd

LogFn = Callable[[str], None]

# 中文姓名 → 英文姓名映射（用于匹配加油交易数据）
NAME_MAPPING: dict[str, str] = {
    '王洋': 'YANG WANG',
    '朱海山': 'HAISHAN ZHU',
    '韩境烨': 'JINGYE HAN',
    '安博': 'BO AN',
    '邢旭': 'XU XING',
    '李健': 'JIAN LI',
    '刘增榕': 'ZENGRONG LIU',
    '陈磊': 'LEI CHEN',
    '吴波': 'BO WU',
}


def _read_driver_data(file_path: Path, log: LogFn) -> pd.DataFrame:
    log(f'读取司机数据: {file_path.name}')
    df = pd.read_excel(file_path)
    df.columns = ['Driver_Name', 'Work_Hours', 'Tasks', 'Mileage', 'Date_Column']
    summary = df.groupby('Driver_Name').agg({
        'Work_Hours': 'sum',
        'Tasks': 'sum',
        'Mileage': 'sum',
    }).reset_index()
    log(f'  司机数: {len(summary)}')
    return summary


def _read_transaction_data(file_path: Path, log: LogFn) -> pd.DataFrame:
    log(f'读取交易数据: {file_path.name}')
    df = pd.read_excel(file_path)
    if 'SiteState' in df.columns:
        il_df = df[df['SiteState'] == 'IL'].copy()
        log(f'  IL 交易: {len(il_df):,} 条，{il_df["Driver"].nunique()} 位司机')
    else:
        log('  WARNING: 无 SiteState 列，使用全部数据')
        il_df = df.copy()
    if 'TransDate' in il_df.columns:
        il_df['TransDate'] = pd.to_datetime(il_df['TransDate'], format='%m/%d/%y')
    return il_df


def _calc_metrics(hours: float, tasks: float, mileage: float) -> dict:
    task_eff = round(tasks / hours, 2) if hours > 0 else 0.0
    mile_eff = round(mileage / hours, 2) if hours > 0 else 0.0
    balance = round(math.sqrt(task_eff * mile_eff), 2) if task_eff > 0 and mile_eff > 0 else 0.0
    avg_dist = round(mileage / tasks, 2) if tasks > 0 else 0.0
    avg_time = round(hours / tasks, 2) if tasks > 0 else 0.0
    workload = round((tasks * mileage) / hours, 2) if hours > 0 else 0.0
    density = round((tasks + mileage / 10) / 6, 2)
    return {
        'Task_Efficiency': task_eff,
        'Mileage_Efficiency': mile_eff,
        'Efficiency_Balance': balance,
        'Avg_Pickup_Distance': avg_dist,
        'Avg_Time_Per_Task': avg_time,
        'Workload_Index': workload,
        'Daily_Workload_Density': density,
    }


def run_dwa(driver_file: Path, transaction_file: Path, output_dir: Path, log: LogFn) -> Path:
    """
    处理司机周数据分析。
    driver_file: dwd_*.xlsx  |  transaction_file: transaction_*.xlsx
    返回输出 Excel 路径。
    """
    log('=' * 60)
    log('司机周数据分析 (DWA)')
    log('=' * 60)

    driver_data = _read_driver_data(driver_file, log)
    transaction_data = _read_transaction_data(transaction_file, log)

    il_drivers: set[str] = set(transaction_data['Driver'].unique()) if transaction_data is not None else set()
    results = []

    log(f'\n分析 {len(driver_data)} 位司机...')
    for _, row in driver_data.iterrows():
        name = row['Driver_Name']
        hours = float(row['Work_Hours'])
        tasks = float(row['Tasks'])
        mileage = float(row['Mileage'])

        mapped = NAME_MAPPING.get(name)
        fuel_count, fuel_cost, fuel_qty = 0, 0.0, 0.0
        matched_name = 'No Match'

        if mapped and mapped in il_drivers:
            dt = transaction_data[transaction_data['Driver'] == mapped]
            fuel_count = len(dt)
            fuel_cost = round(dt['TotalAmount'].sum(), 2) if 'TotalAmount' in dt.columns else 0.0
            fuel_qty = round(dt['Quantity'].sum(), 2) if 'Quantity' in dt.columns else 0.0
            matched_name = mapped
            log(f'  {name} → {mapped}: {fuel_count} 次加油, ${fuel_cost}')
        else:
            log(f'  {name}: 无匹配')

        metrics = _calc_metrics(hours, tasks, mileage)
        results.append({
            'Driver_Name': name,
            'Work_Hours': round(hours, 2),
            'Tasks': int(tasks),
            'Mileage': mileage,
            'Fuel_Count': fuel_count,
            'Fuel_Cost': fuel_cost,
            **metrics,
            'Matched': matched_name,
        })

    df_result = pd.DataFrame(results)

    df_out = df_result[[
        'Driver_Name', 'Work_Hours', 'Tasks', 'Mileage', 'Fuel_Count', 'Fuel_Cost',
        'Task_Efficiency', 'Mileage_Efficiency', 'Efficiency_Balance',
        'Avg_Pickup_Distance', 'Avg_Time_Per_Task', 'Workload_Index', 'Daily_Workload_Density',
    ]].copy()
    df_out.columns = [
        '司机', '总时长(小时)', '总任务(个)', '总里程(英里)', '加油次数', '加油总耗费($)',
        '任务效率(任务/小时)', '里程效率(英里/小时)', '效率平衡分数(√(任务效率×里程效率))',
        '平均揽收间距(英里/任务)', '每任务时长(小时/任务)',
        '工作负荷指数(任务×里程/小时)', '综合工作密度((任务+里程/10)/6天)',
    ]

    match = re.search(r'dwd_(\w+)', driver_file.stem, re.IGNORECASE)
    date_str = match.group(1) if match else driver_file.stem

    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f'dwa_{date_str}.xlsx'
    df_out.to_excel(output_path, index=False)
    log(f'\n✓ 已保存: {output_path.name}')

    return output_path


def _safe_pct(new_val: float, old_val: float) -> float:
    return round((new_val - old_val) / old_val * 100, 1) if old_val != 0 else 0.0


def _generate_explanation(
    t_old, t_new, t_pct,
    m_old, m_new, m_pct,
    h_old, h_new, h_pct,
    d_old, d_new, d_pct,
    wl_old, wl_new, wl_pct,
    p_old, p_new, p_pct,
) -> str:
    parts = []
    if abs(t_pct) >= 10:
        parts.append(f'任务数{"增加" if t_pct > 0 else "减少"}{abs(t_pct):.1f}%({int(t_old)}→{int(t_new)}个)')
    if abs(d_pct) >= 10:
        parts.append(f'每个任务距离{"增加" if d_pct > 0 else "减少"}{abs(d_pct):.1f}%({d_old:.1f}→{d_new:.1f}英里)')
    if abs(h_pct) >= 5:
        parts.append(f'工作时长{"增加" if h_pct > 0 else "减少"}{abs(h_pct):.1f}%({h_old:.1f}→{h_new:.1f}小时)')
    explanation = '，'.join(parts) + '。' if parts else '各项指标基本稳定。'
    if (d_pct > 10 and wl_pct < -5) or (d_pct < -10 and wl_pct > 5):
        explanation += f'总工作量({int(t_old)}×{int(m_old)})→({int(t_new)}×{int(m_new)})'
        if p_pct < 0:
            explanation += f'减少{abs(p_pct):.1f}%，导致工作负荷从{int(wl_old)}降至{int(wl_new)}。'
        else:
            explanation += f'增加{abs(p_pct):.1f}%，导致工作负荷从{int(wl_old)}升至{int(wl_new)}。'
    return explanation


def run_comparison(prev_file: Path, curr_file: Path, output_dir: Path, log: LogFn) -> Path:
    """
    生成周对周对比报表。
    prev_file / curr_file: run_dwa 的输出文件 (dwa_*.xlsx)
    返回输出 Excel 路径。
    """
    log('=' * 60)
    log('周对周对比报表')
    log('=' * 60)
    log(f'上周: {prev_file.name}')
    log(f'本周: {curr_file.name}')

    df_prev = pd.read_excel(prev_file)
    df_curr = pd.read_excel(curr_file)

    common_drivers = sorted(set(df_prev['司机'].values) & set(df_curr['司机'].values))
    log(f'共同司机: {len(common_drivers)} 位')

    report_data = []
    for driver in common_drivers:
        p = df_prev[df_prev['司机'] == driver].iloc[0]
        c = df_curr[df_curr['司机'] == driver].iloc[0]

        t_old, t_new = float(p['总任务(个)']), float(c['总任务(个)'])
        m_old, m_new = float(p['总里程(英里)']), float(c['总里程(英里)'])
        h_old, h_new = float(p['总时长(小时)']), float(c['总时长(小时)'])
        d_old, d_new = float(p['平均揽收间距(英里/任务)']), float(c['平均揽收间距(英里/任务)'])
        wl_old, wl_new = float(p['工作负荷指数(任务×里程/小时)']), float(c['工作负荷指数(任务×里程/小时)'])

        t_pct = _safe_pct(t_new, t_old)
        m_pct = _safe_pct(m_new, m_old)
        h_pct = _safe_pct(h_new, h_old)
        d_pct = _safe_pct(d_new, d_old)
        wl_pct = _safe_pct(wl_new, wl_old)
        p_old = t_old * m_old
        p_new = t_new * m_new
        p_pct = _safe_pct(p_new, p_old)

        explanation = _generate_explanation(
            t_old, t_new, t_pct,
            m_old, m_new, m_pct,
            h_old, h_new, h_pct,
            d_old, d_new, d_pct,
            wl_old, wl_new, wl_pct,
            p_old, p_new, p_pct,
        )

        report_data.append({
            '司机': driver,
            '上周任务数': int(t_old), '本周任务数': int(t_new), '任务数变化%': t_pct,
            '上周总里程': round(m_old, 0), '本周总里程': round(m_new, 0), '总里程变化%': m_pct,
            '上周工作时长': round(h_old, 1), '本周工作时长': round(h_new, 1), '工作时长变化%': h_pct,
            '上周平均揽收间距': round(d_old, 1), '本周平均揽收间距': round(d_new, 1), '揽收间距变化%': d_pct,
            '上周工作负荷': round(wl_old, 0), '本周工作负荷': round(wl_new, 0), '工作负荷变化%': wl_pct,
            '上周(任务×里程)': int(p_old), '本周(任务×里程)': int(p_new), '(任务×里程)变化%': p_pct,
            '数据解释': explanation,
        })

    df_report = pd.DataFrame(report_data)

    def _date_tag(fn: Path) -> str:
        m = re.search(r'dwa_(\w+)', fn.stem)
        return m.group(1) if m else fn.stem

    date_range = f'{_date_tag(prev_file)}-{_date_tag(curr_file)}'
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f'weekly_comparison_{date_range}.xlsx'

    from openpyxl.styles import PatternFill
    prev_fill = PatternFill(start_color='D6EAF8', end_color='D6EAF8', fill_type='solid')
    curr_fill = PatternFill(start_color='D5F4E6', end_color='D5F4E6', fill_type='solid')
    prev_cols = {'上周任务数', '上周总里程', '上周工作时长', '上周平均揽收间距', '上周工作负荷', '上周(任务×里程)'}
    curr_cols = {'本周任务数', '本周总里程', '本周工作时长', '本周平均揽收间距', '本周工作负荷', '本周(任务×里程)'}

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_report.to_excel(writer, sheet_name='周对周对比', index=False)
        ws = writer.sheets['周对周对比']
        for ci, col_name in enumerate(df_report.columns, start=1):
            if col_name in prev_cols:
                for ri in range(1, len(df_report) + 2):
                    ws.cell(row=ri, column=ci).fill = prev_fill
            elif col_name in curr_cols:
                for ri in range(1, len(df_report) + 2):
                    ws.cell(row=ri, column=ci).fill = curr_fill

    log(f'\n✓ 已保存: {output_path.name}')
    return output_path


def run_weekly_chart(prev_file: Path, curr_file: Path, output_dir: Path, log: LogFn) -> Path:
    """
    读取两周的 DWA Excel，生成四图对比周报 PNG。
    返回输出图片路径。
    """
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    log('=' * 60)
    log('生成周报图表')
    log('=' * 60)
    log(f'上周: {prev_file.name}')
    log(f'本周: {curr_file.name}')

    df_prev = pd.read_excel(prev_file)
    df_curr = pd.read_excel(curr_file)

    common = sorted(set(df_prev['司机']) & set(df_curr['司机']))
    log(f'共同司机: {len(common)} 位')

    df_p = df_prev[df_prev['司机'].isin(common)].sort_values('司机').reset_index(drop=True)
    df_c = df_curr[df_curr['司机'].isin(common)].sort_values('司机').reset_index(drop=True)

    drivers = df_c['司机'].values
    x = range(len(drivers))
    w = 0.35

    plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'Arial Unicode MS', 'DejaVu Sans']
    plt.rcParams['axes.unicode_minus'] = False

    def _week(f: Path) -> str:
        m = re.search(r'dwa_\d{2}(\d{2})(\d{2})', f.stem)
        return f'{m.group(1)}/{m.group(2)}' if m else f.stem

    def _labels(ax, bars, fmt, color='#333333'):
        for b in bars:
            h = b.get_height()
            ax.text(b.get_x() + b.get_width() / 2, h,
                    fmt.format(h), ha='center', va='bottom', fontsize=8, color=color, weight='bold')

    curr_wk, prev_wk = _week(curr_file), _week(prev_file)

    fig = plt.figure(figsize=(18, 12))
    gs = fig.add_gridspec(2, 2, hspace=0.3, wspace=0.3)

    # ── 左上: 总任务数 & 总里程（当前周，双轴）
    ax1 = fig.add_subplot(gs[0, 0])
    ax1r = ax1.twinx()
    b1 = ax1.bar([i - w/2 for i in x], df_c['总任务(个)'], w, color='#3498db', alpha=0.8, label='总任务数（个）')
    b2 = ax1r.bar([i + w/2 for i in x], df_c['总里程(英里)'], w, color='#2ecc71', alpha=0.8, label='总里程（英里）')
    ax1.set_xticks(x); ax1.set_xticklabels(drivers, fontsize=10)
    ax1.set_ylabel('总任务数（个）', fontsize=11, color='#3498db'); ax1.tick_params(axis='y', labelcolor='#3498db')
    ax1r.set_ylabel('总里程（英里）', fontsize=11, color='#2ecc71'); ax1r.tick_params(axis='y', labelcolor='#2ecc71')
    _labels(ax1, b1, '{:.0f}', '#3498db'); _labels(ax1r, b2, '{:.0f}', '#2ecc71')
    ax1.set_title(f'总任务数与总里程）', fontsize=13, weight='bold')
    h1, l1 = ax1.get_legend_handles_labels(); h2, l2 = ax1r.get_legend_handles_labels()
    ax1.legend(h1 + h2, l1 + l2, loc='upper left', fontsize=9)

    # ── 右上: 任务效率 & 里程效率（当前周，双轴）
    ax2 = fig.add_subplot(gs[0, 1])
    ax2r = ax2.twinx()
    b3 = ax2.bar([i - w/2 for i in x], df_c['任务效率(任务/小时)'], w, color='#9b59b6', alpha=0.8, label='任务效率（任务/小时）')
    b4 = ax2r.bar([i + w/2 for i in x], df_c['里程效率(英里/小时)'], w, color='#e74c3c', alpha=0.8, label='里程效率（英里/小时）')
    ax2.set_xticks(x); ax2.set_xticklabels(drivers, fontsize=10)
    ax2.set_ylabel('任务效率（任务/小时）', fontsize=11, color='#9b59b6'); ax2.tick_params(axis='y', labelcolor='#9b59b6')
    ax2r.set_ylabel('里程效率（英里/小时）', fontsize=11, color='#e74c3c'); ax2r.tick_params(axis='y', labelcolor='#e74c3c')
    _labels(ax2, b3, '{:.2f}', '#9b59b6'); _labels(ax2r, b4, '{:.2f}', '#e74c3c')
    ax2.set_title(f'任务效率与里程效率', fontsize=13, weight='bold')
    h1, l1 = ax2.get_legend_handles_labels(); h2, l2 = ax2r.get_legend_handles_labels()
    ax2.legend(h1 + h2, l1 + l2, loc='upper left', fontsize=9)

    # ── 左下: 平均揽收间距对比
    ax3 = fig.add_subplot(gs[1, 0])
    bp = ax3.bar([i - w/2 for i in x], df_p['平均揽收间距(英里/任务)'], w, color='#95a5a6', alpha=0.8, label=f'上周({prev_wk})')
    bc = ax3.bar([i + w/2 for i in x], df_c['平均揽收间距(英里/任务)'], w, color='#3498db', alpha=0.8, label=f'本周({curr_wk})')
    ax3.set_xticks(x); ax3.set_xticklabels(drivers, fontsize=10)
    ax3.set_ylabel('平均揽收间距 (英里/任务)', fontsize=11, weight='bold')
    ax3.set_title('平均揽收间距对比', fontsize=13, weight='bold')
    ax3.legend(fontsize=9); ax3.grid(axis='y', alpha=0.3, linestyle='--')
    _labels(ax3, bp, '{:.1f}'); _labels(ax3, bc, '{:.1f}')

    # ── 右下: 工作负荷指数对比
    ax4 = fig.add_subplot(gs[1, 1])
    bp4 = ax4.bar([i - w/2 for i in x], df_p['工作负荷指数(任务×里程/小时)'], w, color='#95a5a6', alpha=0.8, label=f'上周({prev_wk})')
    bc4 = ax4.bar([i + w/2 for i in x], df_c['工作负荷指数(任务×里程/小时)'], w, color='#3498db', alpha=0.8, label=f'本周({curr_wk})')
    ax4.set_xticks(x); ax4.set_xticklabels(drivers, fontsize=10)
    ax4.set_ylabel('工作负荷指数 (任务×里程/小时)', fontsize=11, weight='bold')
    ax4.set_title('工作负荷指数对比', fontsize=13, weight='bold')
    ax4.legend(fontsize=9); ax4.grid(axis='y', alpha=0.3, linestyle='--')
    _labels(ax4, bp4, '{:.0f}'); _labels(ax4, bc4, '{:.0f}')

    fig.suptitle(f'司机周报 — {curr_wk} 数据概览', fontsize=16, weight='bold', y=0.998)
    plt.tight_layout()

    output_dir.mkdir(parents=True, exist_ok=True)
    date_tag = re.sub(r'^dwa_', '', curr_file.stem)
    output_path = output_dir / f'weekly_report_{date_tag}.png'
    plt.savefig(output_path, dpi=150, bbox_inches='tight', facecolor='white')
    plt.close(fig)
    log(f'✓ 已保存: {output_path.name}')
    return output_path
