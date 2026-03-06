"""Step 1: 更新自打面单统计表(h)"""

from copy import copy
from datetime import datetime, timedelta
from pathlib import Path
from typing import Callable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator


def _copy_style_from_above(ws, row: int, col: int):
    """将上一行同列单元格的样式复制到目标单元格。"""
    if row <= 2:
        return
    src = ws.cell(row=row - 1, column=col)
    dst = ws.cell(row=row, column=col)
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.number_format = src.number_format
    dst.protection = copy(src.protection)

SIGNED_STATUSES = {'准备派送', '派送成功', '派送失败', '派送中'}

# 列映射: T-N → {order_count: col_idx, signed: col_idx} (0-based)
COLUMN_MAP = {
    1:  {'order_count': 1, 'signed': 3},
    2:  {'signed': 4},   3:  {'signed': 5},   4:  {'signed': 6},
    5:  {'signed': 7},   6:  {'signed': 8},   7:  {'signed': 9},
    8:  {'signed': 10},  9:  {'signed': 11},  10: {'signed': 12},
    11: {'signed': 13},  12: {'signed': 14},  13: {'signed': 15},
    14: {'signed': 16},
}


def _get_order_time_range(target: datetime):
    end = target.replace(hour=8, minute=0, second=0, microsecond=0)
    return end - timedelta(days=1), end


def _build_date_row_map(ws) -> dict:
    """一次性扫描工作表，建立 date → row 映射，避免 O(n) 重复扫描。"""
    result = {}
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val is None:
            continue
        if isinstance(val, datetime):
            cell_date = val.replace(hour=0, minute=0, second=0, microsecond=0)
        elif isinstance(val, (int, float)):
            cell_date = datetime(1899, 12, 30) + timedelta(days=int(val))
        elif isinstance(val, str):
            try:
                cell_date = datetime.strptime(val, '%Y-%m-%d')
            except ValueError:
                continue
        else:
            continue
        result[cell_date] = row
    return result


def _calculate_c_column(ws, last_row: int) -> int:
    # C[N] = D[N] + (E[N-1]-D[N-1]) + (F[N-2]-E[N-2]) + ... + (Q[N-13]-P[N-13])
    # D=4, E=5, ..., Q=17
    def get_val(row, col):
        if row < 2:
            return 0
        v = ws.cell(row=row, column=col).value
        return v if isinstance(v, (int, float)) else 0

    result = get_val(last_row, 4)
    for i, col in enumerate(range(5, 18)):
        target_row = last_row - 1 - i
        if target_row < 2:
            break
        result += get_val(target_row, col) - get_val(target_row, col - 1)
    return int(result)


def run(today: datetime, data_dir: Path, output_dir: Path, log: Callable) -> dict:
    orders_dir = data_dir / 'input' / 'self_printed'
    xlsx_path = data_dir / 'self_printed_orders.xlsx'

    # 加载订单
    order_files = list(orders_dir.glob('order_*.xlsx'))
    if not order_files:
        raise FileNotFoundError(f'未找到订单文件于 {orders_dir}')

    dfs = []
    for f in order_files:
        log(f'读取: {f.name}')
        df = pd.read_excel(f, usecols=['下单时间', '状态'])
        dfs.append(df)

    orders_df = pd.concat(dfs, ignore_index=True)
    orders_df['下单时间'] = pd.to_datetime(orders_df['下单时间'], format='%m/%d/%Y %H:%M:%S')
    log(f'共加载 {len(orders_df)} 条订单')

    wb = load_workbook(xlsx_path)
    ws = wb.active
    date_row_map = _build_date_row_map(ws)
    updates = []

    for days_ago in range(1, 15):
        target = today - timedelta(days=days_ago)
        t = target.replace(hour=0, minute=0, second=0, microsecond=0)
        row = date_row_map.get(t)

        if row is None:
            if days_ago == 1:
                row = ws.max_row + 1
                prev_row = row - 1
                ws.cell(row=row, column=1, value=target.strftime('%Y-%m-%d'))
                _copy_style_from_above(ws, row, 1)
                date_row_map[t] = row  # 同步更新 map
                log(f'新增行 {row}: {target.strftime("%Y-%m-%d")}')
                prev_c = ws.cell(row=prev_row, column=3)
                if prev_c.value and isinstance(prev_c.value, str) and prev_c.value.startswith('='):
                    ws.cell(row=row, column=3,
                            value=Translator(prev_c.value, origin=f'C{prev_row}').translate_formula(f'C{row}'))
            else:
                log(f'警告: 找不到 {target.strftime("%m/%d")} (T-{days_ago})，跳过')
                continue

        start_t, end_t = _get_order_time_range(target)
        mask = (orders_df['下单时间'] > start_t) & (orders_df['下单时间'] <= end_t)
        day_orders = orders_df[mask]
        order_count = len(day_orders)
        signed_count = int(day_orders['状态'].isin(SIGNED_STATUSES).sum())

        col_config = COLUMN_MAP.get(days_ago, {})

        if 'order_count' in col_config:
            col = col_config['order_count'] + 1
            cell = ws.cell(row=row, column=col)
            if cell.value is None:
                cell.value = order_count
                _copy_style_from_above(ws, row, col)
                updates.append(f'{target.strftime("%m/%d")} B(下单量): {order_count}')

        if 'signed' in col_config:
            col = col_config['signed'] + 1
            cell = ws.cell(row=row, column=col)
            if cell.value is None:
                cell.value = signed_count
                _copy_style_from_above(ws, row, col)
                col_name = chr(ord('A') + col_config['signed'])
                updates.append(f'{target.strftime("%m/%d")} {col_name}(签入): {signed_count}')

    # 计算操作量（在保存前用已有 ws，避免重复开文件）
    sp_op_count = _calculate_c_column(ws, ws.max_row)
    log(f'自打面单操作量: {sp_op_count}')

    if updates:
        wb.save(xlsx_path)
        log(f'已更新 {len(updates)} 个单元格')
        for u in updates:
            log(f'  {u}')
    else:
        log('所有目标单元格已有数据，无需更新')

    return {'sp_operation_count': sp_op_count}
