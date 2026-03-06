"""Step 0: 整合订单(a) + 更新揽收数据表(b)"""

import glob
import re
import sys
from copy import copy
from datetime import datetime, timedelta
from pathlib import Path
from typing import Callable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator


def _copy_style_from_above(ws, row: int, col: int):
    """将上一行同列单元格的样式复制到目标单元格。"""
    if row <= 2:
        return
    src = ws.cell(row=row - 1, column=col)
    dst = ws.cell(row=row, column=col)
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.number_format = src.number_format
    dst.protection = copy(src.protection)

# O列用到的签入列（1-based）: C=3,F=6,I=9,L=12,P=16,T=20,U=21,V=22,W=23,X=24,Y=25,Z=26,AA=27
_SIGN_COLS = [3, 6, 9, 12, 16, 20, 21, 22, 23, 24, 25, 26, 27]

SIGNED_IN_STATUS = ['准备派送', '派送成功', '派送失败', '派送中']
PENDING_STATUS = ['已下单']


# ── a: 整合订单 ──────────────────────────────────────────────────────────────

def _parse_date_range(date_range: str, year: int = 2026):
    start_str, end_str = date_range.split('-')
    start = datetime(year, int(start_str[:2]), int(start_str[2:]), 8) - timedelta(days=1)
    end = datetime(year, int(end_str[:2]), int(end_str[2:]), 8)
    return start, end


def _check_date_coverage(df: pd.DataFrame, start: datetime, end: datetime) -> list[str]:
    df['_dt'] = pd.to_datetime(df['下单时间'], format='%m/%d/%Y %H:%M:%S')
    filtered = df[(df['_dt'] >= start) & (df['_dt'] < end)]
    existing = set(filtered['_dt'].dt.date)
    all_dates = set()
    cur = start.date()
    while cur < end.date():
        all_dates.add(cur)
        cur += timedelta(days=1)
    missing = sorted(all_dates - existing)
    return [d.strftime('%m/%d/%Y') for d in missing]


def _merge_orders(pre_dir: Path, log: Callable) -> pd.DataFrame:
    files = list(pre_dir.glob('order_*.xlsx'))
    if not files:
        raise FileNotFoundError(f'未找到 order_*.xlsx 于 {pre_dir}')
    log(f'找到 {len(files)} 个订单文件')
    dfs = []
    for f in files:
        log(f'  读取: {f.name}')
        dfs.append(pd.read_excel(f))
    merged = pd.concat(dfs, ignore_index=True)
    log(f'整合完成，共 {len(merged)} 行')
    return merged


# ── b: 更新揽收数据表 ─────────────────────────────────────────────────────────

def _filter_orders_by_date(df: pd.DataFrame, target: datetime) -> pd.DataFrame:
    start = (target - timedelta(days=1)).replace(hour=8, minute=0, second=0, microsecond=0)
    end = target.replace(hour=8, minute=0, second=0, microsecond=0)
    return df[(df['下单时间'] >= start) & (df['下单时间'] < end)]


def _count_stats(df: pd.DataFrame) -> dict:
    if df.empty:
        return {'下单量': 0, '签入': 0, '已下单': 0}
    total = len(df)
    active = df[df['状态'] != '已取消']
    return {
        '下单量': total,
        '签入': len(active[active['状态'].isin(SIGNED_IN_STATUS)]),
        '已下单': len(active[active['状态'].isin(PENDING_STATUS)]),
    }


def _get_row_by_date(df: pd.DataFrame, target: datetime) -> int:
    t = target.replace(hour=0, minute=0, second=0, microsecond=0)
    matches = df[df['日期'] == t]
    return matches.index[0] if not matches.empty else -1


def _calculate_o_column(ws, last_row: int) -> int:
    def get_val(row, col):
        if row < 2:
            return 0
        v = ws.cell(row=row, column=col).value
        return v if isinstance(v, (int, float)) else 0

    result = get_val(last_row, _SIGN_COLS[0])
    for i in range(1, len(_SIGN_COLS)):
        result += get_val(last_row - i, _SIGN_COLS[i]) - get_val(last_row - i, _SIGN_COLS[i - 1])
    return int(result)


def _extend_o_formula(ws, last_row: int):
    cell = ws.cell(row=last_row, column=15)
    if cell.value is None:
        prev = ws.cell(row=last_row - 1, column=15)
        if prev.value and isinstance(prev.value, str) and prev.value.startswith('='):
            cell.value = Translator(prev.value, origin=f'O{last_row - 1}').translate_formula(f'O{last_row}')
            _copy_style_from_above(ws, last_row, 15)


# ── main ─────────────────────────────────────────────────────────────────────

def run(today: datetime, data_dir: Path, output_dir: Path, log: Callable) -> dict:
    pre_dir = data_dir / 'input' / 'temu_orders'
    data_file = data_dir / 'gofo_pickup_data.xlsx'
    sheet_name = '下单'

    # 日期区间 T-13 到 T-1
    date_range = f"{(today - timedelta(days=13)).strftime('%m%d')}-{(today - timedelta(days=1)).strftime('%m%d')}"
    log(f'日期区间: {date_range}')
    start_date, end_date = _parse_date_range(date_range)

    # a: 整合
    merged_df = _merge_orders(pre_dir, log)

    log('检查日期完整性...')
    missing = _check_date_coverage(merged_df, start_date, end_date)
    if missing:
        raise RuntimeError(f'缺失日期: {", ".join(missing)}')
    log('日期完整性检查通过')

    if '下单时间_parsed' in merged_df.columns:
        merged_df = merged_df.drop(columns=['下单时间_parsed'])
    if '_dt' in merged_df.columns:
        merged_df = merged_df.drop(columns=['_dt'])

    orig_start = start_date + timedelta(days=1)
    csv_name = f"order_merged_{orig_start.strftime('%y%m%d')}-{end_date.strftime('%y%m%d')}.csv"
    csv_path = data_dir / 'input' / csv_name
    merged_df.to_csv(csv_path, index=False, encoding='utf-8-sig')
    log(f'已保存: {csv_name}')

    # b: 去重
    orders_df = merged_df.drop_duplicates(subset=['系统单号'], keep='first')
    orders_df['下单时间'] = pd.to_datetime(orders_df['下单时间'])

    # 更新 gofo_pickup_data.xlsx
    data_df = pd.read_excel(data_file, sheet_name=sheet_name)
    data_df['日期'] = pd.to_datetime(data_df['日期'])
    wb = load_workbook(data_file)
    ws = wb[sheet_name]
    col_map = {cell.value: cell.column for cell in ws[1]}

    updates = [
        (1,  ['下单量', '当日签入', '当日签入率', '当日已下单']),
        (2,  ['次日签入', '次日签入率', '次日已下单']),
        (3,  ['第三日签入', '第三日签入率', '第三日已下单']),
        (4,  ['第四日签入', '第四日签入率', '第四日已下单']),
        (5,  ['最新签入', '最新签入率']),
        (6,  ['第六日签入']),  (7,  ['第七日签入']),  (8,  ['第八日签入']),
        (9,  ['第九日签入']),  (10, ['第十日签入']),  (11, ['第十一日签入']),
        (12, ['第十二日签入']), (13, ['第十三日签入']),
    ]

    for offset, cols in updates:
        target = today - timedelta(days=offset)
        row_idx = _get_row_by_date(data_df, target)

        if row_idx == -1:
            if offset == 1:
                new_row = {col: None for col in data_df.columns}
                new_row['日期'] = target.replace(hour=0, minute=0, second=0, microsecond=0)
                row_idx = len(data_df)
                data_df.loc[row_idx] = new_row
                excel_row = row_idx + 2
                date_col = col_map.get('日期', 1)
                ws.cell(row=excel_row, column=date_col).value = target.strftime('%Y-%m-%d')
                _copy_style_from_above(ws, excel_row, date_col)
                log(f'新增日期行: {target.strftime("%Y-%m-%d")}')
            else:
                log(f'跳过 {target.strftime("%Y-%m-%d")}: 不在表中')
                continue

        excel_row = row_idx + 2
        filtered = _filter_orders_by_date(orders_df, target)
        stats = _count_stats(filtered)

        if stats['下单量'] == 0:
            log(f'跳过 {target.strftime("%Y-%m-%d")}: 无订单')
            continue

        order_count = data_df.at[row_idx, '下单量'] if pd.notna(data_df.at[row_idx, '下单量']) else stats['下单量']

        updated = False
        for col_name in cols:
            if col_name not in col_map:
                continue
            cell = ws.cell(row=excel_row, column=col_map[col_name])
            if cell.value is not None:
                continue
            updated = True
            if col_name == '下单量':
                cell.value = stats['下单量']
            elif '签入率' in col_name:
                cell.value = round(stats['签入'] / order_count * 100, 1) if order_count > 0 else 0
            elif '已下单' in col_name:
                cell.value = stats['已下单']
            elif '签入' in col_name:
                cell.value = stats['签入']
            _copy_style_from_above(ws, excel_row, col_map[col_name])

        if updated:
            log(f'更新 {target.strftime("%Y-%m-%d")}: 下单量={stats["下单量"]}, 签入={stats["签入"]}')

    # 延伸 O 列公式并保存
    _extend_o_formula(ws, ws.max_row)
    wb.save(data_file)
    log(f'已保存 {data_file.name}')

    # 读取 T-1 数据
    last_row = ws.max_row
    t1_order_total = ws.cell(row=last_row, column=2).value or 0
    t1_signed_in   = ws.cell(row=last_row, column=3).value or 0
    t1_signin_rate = ws.cell(row=last_row, column=4).value or 0
    t1_pending     = ws.cell(row=last_row, column=5).value or 0
    tubt_op_signin = _calculate_o_column(ws, last_row)

    log(f'T-1 下单量={t1_order_total}, 签入={t1_signed_in}, 签入率={t1_signin_rate}%, 操作签入量={tubt_op_signin}')

    return {
        't1_order_total': t1_order_total,
        't1_signed_in': t1_signed_in,
        't1_signin_rate': t1_signin_rate,
        't1_pending': t1_pending,
        'tubt_operation_signin': tubt_op_signin,
    }
