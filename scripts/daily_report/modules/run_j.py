"""Step 5: 预约达成率与揽收占比(j)"""

from copy import copy
from datetime import datetime, timedelta
from pathlib import Path
from typing import Callable

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook

from utils.chinese_font import setup_chinese_font


def _copy_style_from_above(ws, row: int, col: int):
    """将上一行同列单元格的样式复制到目标单元格。"""
    if row <= 2:
        return
    src = ws.cell(row=row - 1, column=col)
    dst = ws.cell(row=row, column=col)
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.number_format = src.number_format
    dst.protection = copy(src.protection)


def run(today: datetime, data_dir: Path, output_dir: Path, log: Callable) -> dict:
    setup_chinese_font()
    plt.rcParams['axes.unicode_minus'] = False

    last2day = today - timedelta(days=2)
    last2day_str = last2day.strftime('%m-%d')
    last2day_full = last2day.strftime('%Y-%m-%d')

    module_dir = data_dir / 'input' / 'yy_pickup_rate'
    reach_file = module_dir / '揽收达成率_完整数据_data.csv'
    signin_file = module_dir / '目的区域签入量_完整数据_data.csv'

    log(f'读取预约达成率数据 ({last2day_str})')
    df_reach = pd.read_csv(reach_file, encoding='utf-16', sep='\t')
    df_signin = pd.read_csv(signin_file, encoding='utf-16', sep='\t')

    def to_float_percent(x):
        if isinstance(x, str) and '%' in x:
            return float(x.replace('%', '')) / 100
        return float(x) if pd.notnull(x) else 0.0

    df_reach['揽收达成率'] = df_reach['揽收达成率'].apply(to_float_percent)
    df_signin['签入率'] = df_signin['签入率'].apply(to_float_percent)

    total_count = len(df_reach)
    reach_count = int((df_reach['揽收达成率'] == 1.0).sum())
    reach_fail_count = int((df_reach['揽收达成率'] == 0.0).sum())
    reach_percent = round(reach_count / total_count * 100, 1) if total_count > 0 else 0

    signin_count = int((df_signin['签入率'] == 1.0).sum())
    signin_fail_count = int((df_signin['签入率'] == 0.0).sum())
    signin_percent = round(signin_count / reach_count * 100, 1) if reach_count > 0 else 0
    fail_percent = round(100 - signin_percent, 1)

    log(f'揽收达成: {reach_count}/{total_count} ({reach_percent}%), 签入: {signin_count} ({signin_percent}%)')

    # 导出未达成单号供 k 模块使用
    df_reach_zero = df_reach[df_reach['揽收达成率'] == 0.0][['单据号']]
    zero_output = data_dir / 'input' / 'yy_unachieved' / f'{last2day_str}未达成单号.csv'
    df_reach_zero.to_csv(zero_output, index=False, encoding='utf-8-sig')
    log(f'已输出 {len(df_reach_zero)} 条未达成单号到 k/')

    # 饼图
    fig, ax = plt.subplots(figsize=(6, 6))
    ax.pie([signin_percent, fail_percent], labels=['揽收达成', '揽收未达成'],
           autopct='%1.1f%%', startangle=90,
           colors=['#4CAF50', '#FF5722'], textprops={'fontsize': 14})
    ax.set_title('揽收达成与未达成占比', fontsize=18)
    ax.axis('equal')
    plt.tight_layout()
    j_image_pie = output_dir / '揽收达成与未达成占比.png'
    fig.savefig(j_image_pie, dpi=150)
    plt.close(fig)

    # 保存到 gofo_pickup_data.xlsx 预约 sheet
    excel_path = data_dir / 'gofo_pickup_data.xlsx'
    wb = load_workbook(excel_path)
    ws = wb['预约']
    last_row = ws.max_row + 1
    ws.cell(row=last_row, column=1, value=last2day_full)
    ws.cell(row=last_row, column=2, value=total_count)
    ws.cell(row=last_row, column=3, value=reach_count)
    ws.cell(row=last_row, column=4, value=reach_percent / 100)
    ws.cell(row=last_row, column=5, value=signin_count)
    ws.cell(row=last_row, column=6, value=signin_percent / 100)
    ws.cell(row=last_row, column=7, value=reach_fail_count)
    ws.cell(row=last_row, column=8, value=0)
    for col in range(1, 9):
        _copy_style_from_above(ws, last_row, col)
    wb.save(excel_path)
    log('已保存到 gofo_pickup_data.xlsx 预约 sheet')

    return {
        'reach_percent': reach_percent,
        'signin_count': signin_count,
        'signin_percent': signin_percent,
        'signin_fail_count': signin_fail_count,
        'fail_percent': fail_percent,
        'df_reach_zero': len(df_reach_zero),
        'j_image_pie': str(j_image_pie),
    }
