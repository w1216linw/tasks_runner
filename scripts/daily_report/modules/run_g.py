"""Step 4: SHEIN D2D 下单(g)"""

import re
from copy import copy
from datetime import datetime, timedelta
from pathlib import Path
from typing import Callable

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook

from utils.chinese_font import setup_chinese_font

SIGN_IN_STATUS = ['准备派送', '派送成功', '派送失败', '派送中']


def _copy_style_from_above(ws, row: int, col: int):
    """将上一行同列单元格的样式复制到目标单元格。"""
    if row <= 2:
        return
    src = ws.cell(row=row - 1, column=col)
    dst = ws.cell(row=row, column=col)
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.number_format = src.number_format
    dst.protection = copy(src.protection)


def _extract_true_address_number(addr):
    addr = str(addr)
    addr = re.sub(r'\b\d{4,5}\b$', '', addr)
    numbers = re.findall(r'\d+', addr)
    return str(max(map(int, numbers))) if numbers else ''


def _extract_address_lib_number(addr):
    numbers = re.findall(r'\d+', str(addr))
    return numbers[0] if numbers else ''


def _extract_unit(addr):
    m = re.search(r'(APT|#|UNIT|SUITE|ROOM|FL|楼)\s*[\w\-]+', str(addr), re.IGNORECASE)
    return m.group(0).upper().strip() if m else ''


def run(today: datetime, data_dir: Path, output_dir: Path, log: Callable) -> dict:
    setup_chinese_font()
    plt.rcParams['axes.unicode_minus'] = False

    last2day = today - timedelta(days=2)
    last2day_str = last2day.strftime('%m-%d')
    last2day_full = last2day.strftime('%Y-%m-%d')
    g_date_str = last2day_full

    module_dir = data_dir / 'input' / 'shein_d2d'
    sn_file = module_dir / f'SN{last2day_str}.xlsx'
    address_path = data_dir / '地址库.xlsx'
    sn_table_path = data_dir / 'SHEIN-D2D.xlsx'

    log(f'读取: {sn_file.name}')
    sn_orders = pd.read_excel(sn_file, engine='openpyxl')
    address_lib = pd.read_excel(address_path, engine='openpyxl')

    total_sn = len(sn_orders)
    sn_no_cancel = sn_orders[~sn_orders['取消原因'].eq('客户取消发货')]
    cancel_sn = int((sn_orders['取消原因'] == '客户取消发货').sum())
    sign_in_sn = int(sn_no_cancel['状态'].isin(SIGN_IN_STATUS).sum())
    sign_in_sn_percent = round(sign_in_sn / total_sn * 100, 1) if total_sn > 0 else 0

    orders_yxd = sn_orders[
        (sn_orders['最新轨迹'] == '已下单') &
        (~sn_orders['取消原因'].eq('客户取消发货'))
    ].copy()
    orders_yxd['门牌号'] = orders_yxd['发件人详细地址'].apply(_extract_true_address_number)
    orders_yxd['单元号'] = orders_yxd['发件人详细地址'].apply(_extract_unit)
    orders_yxd['地址单位'] = orders_yxd.apply(
        lambda row: f"{row['门牌号']} {row['单元号']}" if row['门牌号'] == '1111' and row['单元号'] else row['门牌号'],
        axis=1
    )
    total_yxd = len(orders_yxd)

    log(f'SN总单: {total_sn}, 取消: {cancel_sn}, 签入: {sign_in_sn} ({sign_in_sn_percent}%), 未达成: {total_yxd}')

    # 图一：触达饼图
    sn_orders['门牌号'] = sn_orders['发件人详细地址'].apply(_extract_true_address_number)
    address_lib['门牌号'] = address_lib['实际主地址'].apply(_extract_address_lib_number)
    sn_seen = set(address_lib['门牌号'].dropna().astype(str))
    sn_orders['是否触达'] = sn_orders['门牌号'].astype(str).apply(lambda x: '已触达' if x in sn_seen else '未触达')
    sn_count = sn_orders['是否触达'].value_counts()

    fig1, ax1 = plt.subplots(figsize=(4, 4))
    ax1.pie(sn_count.values, labels=sn_count.index, autopct='%1.1f%%',
            colors=['#4CAF50', '#F44336'][:len(sn_count)], startangle=140,
            explode=[0.05] + [0] * (len(sn_count) - 1))
    ax1.set_title(f'SN-D2D地址触达情况（{g_date_str}）', fontsize=14)
    ax1.axis('equal')
    plt.tight_layout()
    g_image_pie = output_dir / f'SN-D2D地址触达情况（{g_date_str}）.png'
    fig1.savefig(g_image_pie, dpi=150)
    plt.close(fig1)

    # 图二：未达成地址分布
    yxd_unit = orders_yxd['地址单位'].value_counts().reset_index()
    yxd_unit.columns = ['地址单位', '数量']
    yxd_total = yxd_unit['数量'].sum()
    yxd_unit = yxd_unit.nlargest(10, '数量').sort_values(by='数量', ascending=True)

    fig2, ax2 = plt.subplots(figsize=(10, 6))
    bars = ax2.barh(yxd_unit['地址单位'], yxd_unit['数量'])
    for bar, cnt in zip(bars, yxd_unit['数量']):
        ax2.text(bar.get_width() + 1, bar.get_y() + bar.get_height() / 2,
                 f'{cnt}单 ({cnt/yxd_total:.1%})' if yxd_total > 0 else f'{cnt}单',
                 va='center', fontsize=9)
    ax2.set_title(f'SN-D2D仍为已下单的前十地址分布 {g_date_str}', fontsize=14)
    ax2.set_xlabel('订单数量')
    plt.tight_layout()
    g_image_bar = output_dir / f'SN-D2D仍为已下单的前十地址分布 {g_date_str}.png'
    fig2.savefig(g_image_bar, dpi=150)
    plt.close(fig2)

    # 保存到 SHEIN-D2D.xlsx
    wb = load_workbook(sn_table_path)
    ws = wb['Sheet1']
    last_row = ws.max_row + 1
    ws.cell(row=last_row, column=1, value=last2day_full)
    ws.cell(row=last_row, column=2, value=total_sn)
    ws.cell(row=last_row, column=3, value=sign_in_sn)
    ws.cell(row=last_row, column=4, value=sign_in_sn_percent / 100)
    ws.cell(row=last_row, column=5, value=sign_in_sn)
    ws.cell(row=last_row, column=6, value=sign_in_sn_percent / 100)
    for col in range(1, 7):
        _copy_style_from_above(ws, last_row, col)
    wb.save(sn_table_path)
    log('已保存到 SHEIN-D2D.xlsx')

    return {
        'total_sn': total_sn,
        'cancel_sn': cancel_sn,
        'sign_in_sn': sign_in_sn,
        'sign_in_sn_percent': sign_in_sn_percent,
        'total_yxd': total_yxd,
        'g_date_str': g_date_str,
        'g_image_pie': str(g_image_pie),
        'g_image_bar': str(g_image_bar),
    }
