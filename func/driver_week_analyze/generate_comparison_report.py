#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成周对周对比报表
"""

import pandas as pd
import argparse
import sys
import os
import warnings
warnings.filterwarnings('ignore')

# Set environment for proper Unicode handling
def setup_console_encoding():
    """Setup console encoding for Windows compatibility"""
    try:
        if sys.platform.startswith('win'):
            os.environ['PYTHONIOENCODING'] = 'utf-8'
            # For Windows compatibility
            import locale
            try:
                locale.setlocale(locale.LC_ALL, 'C.UTF-8')
            except:
                pass
    except Exception:
        pass

setup_console_encoding()

def generate_explanation(driver, t_old, t_new, t_pct, m_old, m_new, m_pct,
                        h_old, h_new, h_pct, d_old, d_new, d_pct,
                        wl_old, wl_new, wl_pct, p_old, p_new, p_pct):
    """生成数据解释"""

    parts = []

    # 任务数变化
    if abs(t_pct) >= 10:
        if t_pct > 0:
            parts.append(f'任务数增加{abs(t_pct):.1f}%({int(t_old)}→{int(t_new)}个)')
        else:
            parts.append(f'任务数减少{abs(t_pct):.1f}%({int(t_old)}→{int(t_new)}个)')

    # 揽收间距变化
    if abs(d_pct) >= 10:
        if d_pct > 0:
            parts.append(f'每个任务距离增加{abs(d_pct):.1f}%({d_old:.1f}→{d_new:.1f}英里)')
        else:
            parts.append(f'每个任务距离减少{abs(d_pct):.1f}%({d_old:.1f}→{d_new:.1f}英里)')

    # 工作时长变化
    if abs(h_pct) >= 5:
        if h_pct > 0:
            parts.append(f'工作时长增加{abs(h_pct):.1f}%({h_old:.1f}→{h_new:.1f}小时)')
        else:
            parts.append(f'工作时长减少{abs(h_pct):.1f}%({h_old:.1f}→{h_new:.1f}小时)')

    # 总工作量变化 - 关键解释
    if len(parts) > 0:
        explanation = '，'.join(parts) + '。'
    else:
        explanation = '各项指标基本稳定。'

    # 特殊情况解释：揽收间距和工作负荷反向变化
    if (d_pct > 10 and wl_pct < -5) or (d_pct < -10 and wl_pct > 5):
        explanation += f'总工作量({int(t_old)}×{int(m_old)})→({int(t_new)}×{int(m_new)})'
        if p_pct < 0:
            explanation += f'减少{abs(p_pct):.1f}%，导致工作负荷从{int(wl_old)}降至{int(wl_new)}。'
        else:
            explanation += f'增加{abs(p_pct):.1f}%，导致工作负荷从{int(wl_old)}升至{int(wl_new)}。'

    return explanation

def parse_arguments():
    """
    Parse command line arguments
    """
    parser = argparse.ArgumentParser(
        description='周对周对比报表生成程序',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python generate_comparison_report.py -f dwa_1020 dwa_1027
  python generate_comparison_report.py -f dwa_1020.xlsx dwa_1027.xlsx
  python generate_comparison_report.py -f dwa_1020 dwa_1027 -o comparison_1020_1027.xlsx
        """
    )

    parser.add_argument('-f', '--files', nargs=2, required=True,
                        metavar=('PREV_WEEK', 'CURR_WEEK'),
                        help='上周和本周的DWA分析文件（支持带.xlsx或不带扩展名）')

    parser.add_argument('-o', '--output', default=None,
                        help='输出文件名（默认：自动生成为 weekly_comparison_PREV_CURR.xlsx）')

    return parser.parse_args()

def extract_date_from_filename(filename):
    """
    从文件名中提取日期部分
    例如: 'dwa_260105.xlsx' -> '260105'
         'dwa_1020.xlsx' -> '1020'
         'dwa_1020' -> '1020'
    """
    import re
    base_name = os.path.splitext(os.path.basename(filename))[0]
    # 提取dwa_后面的日期部分（支持4-6位数字）
    match = re.search(r'dwa_(\d{4,6})', base_name)
    if match:
        return match.group(1)
    # 如果没有匹配，尝试提取纯数字
    match = re.search(r'(\d{4,6})', base_name)
    if match:
        return match.group(1)
    return base_name

def main():
    # 解析命令行参数
    args = parse_arguments()

    # 处理文件名（如果没有.xlsx扩展名则添加）
    prev_file = args.files[0] if args.files[0].endswith('.xlsx') else f"{args.files[0]}.xlsx"
    curr_file = args.files[1] if args.files[1].endswith('.xlsx') else f"{args.files[1]}.xlsx"

    # 生成输出文件名
    if args.output is None:
        prev_date = extract_date_from_filename(prev_file)
        curr_date = extract_date_from_filename(curr_file)
        output_file = f"weekly_comparison_{prev_date}-{curr_date}.xlsx"
    else:
        output_file = args.output

    print("=== 周对周对比报表生成 ===")
    print(f"上周数据: {prev_file}")
    print(f"本周数据: {curr_file}")
    print(f"输出文件: {output_file}")

    # 检查文件是否存在
    if not os.path.exists(prev_file):
        print(f"错误: 找不到文件 '{prev_file}'")
        sys.exit(1)
    if not os.path.exists(curr_file):
        print(f"错误: 找不到文件 '{curr_file}'")
        sys.exit(1)

    print("\n读取数据文件...")
    # 读取两周数据
    df_prev = pd.read_excel(prev_file)
    df_curr = pd.read_excel(curr_file)

    # 找到共同司机
    common_drivers = sorted(set(df_prev['司机'].values) & set(df_curr['司机'].values))

    print(f"找到 {len(common_drivers)} 个共同司机")
    print("\n生成对比分析...")

    # 创建对比表格数据
    report_data = []

    for driver in common_drivers:
        data_prev = df_prev[df_prev['司机'] == driver].iloc[0]
        data_curr = df_curr[df_curr['司机'] == driver].iloc[0]

        # 提取数据
        tasks_old = data_prev['总任务(个)']
        tasks_new = data_curr['总任务(个)']
        miles_old = data_prev['总里程(英里)']
        miles_new = data_curr['总里程(英里)']
        hours_old = data_prev['总时长(小时)']
        hours_new = data_curr['总时长(小时)']
        task_eff_old = data_prev['任务效率(任务/小时)']
        task_eff_new = data_curr['任务效率(任务/小时)']
        avg_dist_old = data_prev['平均揽收间距(英里/任务)']
        avg_dist_new = data_curr['平均揽收间距(英里/任务)']
        wl_old = data_prev['工作负荷指数(任务×里程/小时)']
        wl_new = data_curr['工作负荷指数(任务×里程/小时)']

        # 计算变化百分比
        tasks_pct = ((tasks_new - tasks_old) / tasks_old * 100) if tasks_old > 0 else 0
        miles_pct = ((miles_new - miles_old) / miles_old * 100) if miles_old > 0 else 0
        hours_pct = ((hours_new - hours_old) / hours_old * 100) if hours_old > 0 else 0
        task_eff_pct = ((task_eff_new - task_eff_old) / task_eff_old * 100) if task_eff_old > 0 else 0
        avg_dist_pct = ((avg_dist_new - avg_dist_old) / avg_dist_old * 100) if avg_dist_old > 0 else 0
        wl_pct = ((wl_new - wl_old) / wl_old * 100) if wl_old > 0 else 0

        # 计算乘积
        product_old = tasks_old * miles_old
        product_new = tasks_new * miles_new
        product_pct = ((product_new - product_old) / product_old * 100) if product_old > 0 else 0

        # 生成解释
        explanation = generate_explanation(
            driver, tasks_old, tasks_new, tasks_pct,
            miles_old, miles_new, miles_pct,
            hours_old, hours_new, hours_pct,
            avg_dist_old, avg_dist_new, avg_dist_pct,
            wl_old, wl_new, wl_pct,
            product_old, product_new, product_pct
        )

        report_data.append({
            '司机': driver,
            '上周任务数': int(tasks_old),
            '本周任务数': int(tasks_new),
            '任务数变化%': round(tasks_pct, 1),
            '上周总里程': round(miles_old, 0),
            '本周总里程': round(miles_new, 0),
            '总里程变化%': round(miles_pct, 1),
            '上周工作时长': round(hours_old, 1),
            '本周工作时长': round(hours_new, 1),
            '工作时长变化%': round(hours_pct, 1),
            '上周平均揽收间距': round(avg_dist_old, 1),
            '本周平均揽收间距': round(avg_dist_new, 1),
            '揽收间距变化%': round(avg_dist_pct, 1),
            '上周工作负荷': round(wl_old, 0),
            '本周工作负荷': round(wl_new, 0),
            '工作负荷变化%': round(wl_pct, 1),
            '上周(任务×里程)': int(product_old),
            '本周(任务×里程)': int(product_new),
            '(任务×里程)变化%': round(product_pct, 1),
            '数据解释': explanation
        })

    # 创建DataFrame
    df_report = pd.DataFrame(report_data)

    # 保存到Excel并添加颜色格式
    print(f"\n保存结果...")
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_report.to_excel(writer, sheet_name='周对周对比', index=False)

        # 获取worksheet对象
        worksheet = writer.sheets['周对周对比']

        # 导入openpyxl的样式模块
        from openpyxl.styles import PatternFill

        # 定义颜色：上周用浅蓝色，本周用浅绿色
        prev_week_fill = PatternFill(start_color='D6EAF8', end_color='D6EAF8', fill_type='solid')  # 浅蓝色
        curr_week_fill = PatternFill(start_color='D5F4E6', end_color='D5F4E6', fill_type='solid')  # 浅绿色

        # 定义上周和本周的列（基于列名）
        prev_week_cols = ['上周任务数', '上周总里程', '上周工作时长', '上周平均揽收间距', '上周工作负荷', '上周(任务×里程)']
        curr_week_cols = ['本周任务数', '本周总里程', '本周工作时长', '本周平均揽收间距', '本周工作负荷', '本周(任务×里程)']

        # 找到这些列的索引（Excel列号从1开始，加1是因为有表头）
        header_row = 1
        for col_idx, col_name in enumerate(df_report.columns, start=1):
            # 给表头行和所有数据行添加颜色
            if col_name in prev_week_cols:
                for row_idx in range(1, len(df_report) + 2):  # +2 包括表头
                    worksheet.cell(row=row_idx, column=col_idx).fill = prev_week_fill
            elif col_name in curr_week_cols:
                for row_idx in range(1, len(df_report) + 2):
                    worksheet.cell(row=row_idx, column=col_idx).fill = curr_week_fill

    print(f'\n✅ 对比报表已保存到: {output_file}')
    print(f'\n数据预览:\n')
    print(df_report[['司机', '任务数变化%', '揽收间距变化%', '工作负荷变化%', '数据解释']].to_string(index=False))

    print(f"\n完成! 共分析 {len(common_drivers)} 个司机的周对周变化")

if __name__ == '__main__':
    main()
